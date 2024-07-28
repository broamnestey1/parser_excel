import openpyxl as op

data = op.load_workbook("full_data.xlsx")
sheet_1 = data.sheetnames[0]
sheet_data = data[sheet_1]

found_RD_full = {}
found_RD_not_full = {}

def full_data():

    found_flag = False  # Для отслеживания найденного столбца Названия РД

    shifr_row = 0  # Инициализация строки для Шифра работ
    shifr_col = 0  # Инициализация столбца для Шифра работ
    rd_row = 0  # Инициализация строки для Названия РД
    rd_col = 0  # Инициализация столбца для Названия РД


    for row in sheet_data.iter_rows():
        for cell in row:
            if cell.value and cell.value.replace(" ", "") == 'Шифр':
                print("Столбец Шифр найден!")
                shifr_row = cell.row
                shifr_col = cell.column
                #print(f"Шифр найден: строка={shifr_row}, столбец={shifr_col}")
                #print("_______________________")
                found_flag = True

            if cell.value and cell.value.replace(" ", "") == 'НазваниеРД':
                print("Столбец с названием РД найден!")
                rd_row = cell.row
                rd_col = cell.column
                #print(f"Строка = {rd_row}")
                #print(f"Столбец = {rd_col}")
                found_flag = True
                break
        if found_flag:
            break

    if not found_flag:
        print("Столбец не найден! Требуется корректировка документа в столбце!")
    else:
        rd_values = [cell[0].value for cell in sheet_data.iter_rows(min_row=rd_row+1, min_col=rd_col, max_col=rd_col)]
        shifr_values = [cell[0].value for cell in sheet_data.iter_rows(min_row=rd_row+1, min_col=shifr_col, max_col=shifr_col)]

        for i in range(len(rd_values)):
            rd_value = rd_values[i]
            if i < len(shifr_values):
                shifr_value = shifr_values[i]
            else:
                shifr_value = "" 
            found_RD_full[rd_value] = shifr_value

        # Вывод результата
        #for rd_value, shifr_value in found_RD.items():
        #    print(f"{rd_value}: {shifr_value}")

        print("Конечный словарь:", found_RD_full)


def not_full_data():
    data_not_full = op.load_workbook("not_full_data.xlsx")
    sheet_1_not = data_not_full.sheetnames[0]
    sheet_data_not = data_not_full[sheet_1_not]


    found_flag_not = False  # Для отслеживания найденного столбца Названия РД
    shifr_row_not = 0
    shifr_col_not = 0  # Инициализация столбца для Шифра работ
    rd_row_not = 0  # Инициализация строки для Названия РД
    rd_col_not = 0  # Инициализация столбца для Названия РД

    for row in sheet_data_not.iter_rows():
        for cell in row:
            if cell.value and cell.value.replace(" ", "") == 'Шифр':
                print("Столбец Шифр найден!")
                shifr_row_not = cell.row
                shifr_col_not = cell.column
                #print(f"Шифр найден: строка={shifr_row}, столбец={shifr_col}")
                #print("_______________________")
                found_flag_not = True

            if cell.value and cell.value.replace(" ", "") == 'НазваниеРД':
                print("Столбец с названием РД найден!")
                rd_row_not = cell.row
                rd_col_not = cell.column
                #print(f"Строка = {rd_row}")
                #print(f"Столбец = {rd_col}")
                found_flag_not = True
                break
        if found_flag_not:
            break

    if not found_flag_not:
        print("Столбец не найден! Требуется корректировка документа в столбце!")
    else:
        rd_values_not = [cell[0].value for cell in sheet_data_not.iter_rows(min_row=rd_row_not+1, min_col=rd_col_not, max_col=rd_col_not)]
        shifr_values_not = [cell[0].value for cell in sheet_data_not.iter_rows(min_row=rd_row_not+1, min_col=shifr_col_not, max_col=shifr_col_not)]

        for i in range(len(rd_values_not)):
            rd_value_not = rd_values_not[i]
            if i < len(shifr_values_not):
                shifr_value_not = shifr_values_not[i]
            else:
                shifr_value_not = "" 
            found_RD_not_full[rd_value_not] = shifr_value_not

        # Вывод результата
        #for rd_value, shifr_value in found_RD.items():
        #    print(f"{rd_value}: {shifr_value}")

        print("Конечный словарь:", found_RD_not_full)


full_data()
not_full_data()







def filter_none_keys(d):
    return {k: v for k, v in d.items() if k is not None}

def compare_dicts(dict1, dict2):
    dict1 = filter_none_keys(dict1)

    keys1 = set(dict1.keys())
    keys2 = set(dict2.keys())

    common_keys = keys1 & keys2

    print("__________________________________________________")

    for key in common_keys:
        if dict1[key] != dict2[key]:
            print(f"Значение для ключа '{key}' отличается: {dict1[key]} != {dict2[key]}")
            dict2[key] = dict1[key]
            print("Обновленный dict2:", dict2)

# Можно также сравнить значения для всех ключей в dict1 и dict2
    #for key in keys1 - keys2:
        #print(f"Ключ '{key}' присутствует только в ИЗНАЧАЛЬНОМ ФАЙЛЕ")

    for key in keys2 - keys1:
        if key is not None:
            print(f"Ключ '{key}' присутствует только в ПРОВЕРЯЕМОМ ФАЙЛЕ")

    #return dict_filtered == dict2

print(compare_dicts(found_RD_full, found_RD_not_full))  # Output: True