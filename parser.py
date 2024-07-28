import openpyxl as op

data_pusto = op.load_workbook('test2.xlsx')
data_true = op.load_workbook('test1.xlsx')

sheet_1_pusto = data_pusto.sheetnames[0]
sheet_1_data = data_pusto[sheet_1_pusto]

sheet_1_true = data_true.sheetnames[0]
sheet_1_true = data_true[sheet_1_true]

row_num = 2
col_num = 1

for row in sheet_1_data.iter_rows(min_row=row_num, min_col=col_num, max_col = 3):
    #print(keys)
    start_row = row
    for cell in start_row:
        #print(cell.value)
        for row_true in sheet_1_true.iter_rows(min_row=2, max_row=2, min_col=col_num, max_col=3):
            start_row_true = row_true
            for cell in start_row_true:
                yacheyka_true = cell.value
                print(yacheyka_true)

        print(cell.value)


for row in sheet_1_true.iter_rows(min_row=row_num, min_col=col_num, max_col = 3):
    #print(keys)
    start_row = row
    for cell in start_row:
        print(cell.value)